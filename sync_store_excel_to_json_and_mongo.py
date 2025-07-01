import os
import json
import openpyxl
from pymongo import MongoClient

EXCEL_PATH = os.path.join('excel_data', 'hookah_store_products.xlsx')
JSON_PATH = os.path.join('mongodb_data', 'store_products.json')

# --- إعداد بيانات الاتصال بمونجو ---
MONGO_URI = 'mongodb://localhost:27017/'  # عدلها إذا كان لديك بيانات اتصال مختلفة
DB_NAME = 'hookah_shop'                   # عدلها إذا كان اسم قاعدة البيانات مختلف
COLLECTION_NAME = 'store_products'        # عدلها إذا كان اسم الكوليكشن مختلف

def rename_excel_column(file_path, old_name, new_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for cell in ws[1]:
        if cell.value == old_name:
            cell.value = new_name
            wb.save(file_path)
            print(f"[Excel] Renamed column '{old_name}' to '{new_name}' in {file_path}")
            return True
    print(f"[Excel] Column '{old_name}' not found in {file_path}")
    return False

def sync_excel_to_json_and_mongo():
    # 1. Rename column in Excel if needed
    rename_excel_column(EXCEL_PATH, 'quantity', 'carton_count')

    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        return
    if not os.path.exists(JSON_PATH):
        print(f"JSON file not found: {JSON_PATH}")
        return
    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    excel_data = {}
    for row in ws.iter_rows(min_row=2):
        id_idx = headers.get('id')
        c_idx = headers.get('carton_count')
        if id_idx and c_idx:
            item_id = str(row[id_idx-1].value)
            carton_count = row[c_idx-1].value
            excel_data[item_id] = carton_count
    # Load JSON
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    updated_json = 0
    for item in data:
        item_id = str(item.get('id'))
        if item_id in excel_data:
            excel_val = excel_data[item_id]
            if excel_val is not None and item.get('carton_count') != excel_val:
                item['carton_count'] = excel_val
                updated_json += 1
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[Sync] Updated {updated_json} items in {JSON_PATH} from Excel.")

    # --- تحديث MongoDB ---
    try:
        client = MongoClient(MONGO_URI)
        db = client[DB_NAME]
        collection = db[COLLECTION_NAME]
        updated_mongo = 0
        for item_id, carton_count in excel_data.items():
            if carton_count is not None:
                result = collection.update_one({'id': item_id}, {'$set': {'carton_count': carton_count}})
                if result.modified_count > 0:
                    updated_mongo += 1
        print(f"[Sync] Updated {updated_mongo} items in MongoDB from Excel.")
    except Exception as e:
        print(f"[MongoDB] Error: {e}")

if __name__ == "__main__":
    sync_excel_to_json_and_mongo() 