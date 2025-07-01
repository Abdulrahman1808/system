import os
import json
import openpyxl

EXCEL_PATH = os.path.join('excel_data', 'hookah_inventory.xlsx')
JSON_PATH = os.path.join('mongodb_data', 'inventory.json')

# --- Sync Excel to JSON ---
def sync_excel_to_json():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        return
    if not os.path.exists(JSON_PATH):
        print(f"JSON file not found: {JSON_PATH}")
        return
    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    excel_data = {}
    for row in ws.iter_rows(min_row=2):
        id_idx = headers.get('id')
        c_idx = headers.get('carton_count')
        u_idx = headers.get('units_per_carton')
        if id_idx:
            item_id = str(row[id_idx-1].value)
            carton_count = row[c_idx-1].value if c_idx else None
            units_per_carton = row[u_idx-1].value if u_idx else None
            excel_data[item_id] = {
                'carton_count': carton_count,
                'units_per_carton': units_per_carton
            }
    # Load JSON
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    updated = 0
    for item in data:
        item_id = str(item.get('id'))
        if item_id in excel_data:
            excel_vals = excel_data[item_id]
            changed = False
            if excel_vals['carton_count'] is not None and item.get('carton_count') != excel_vals['carton_count']:
                item['carton_count'] = excel_vals['carton_count']
                changed = True
            if excel_vals['units_per_carton'] is not None and item.get('units_per_carton') != excel_vals['units_per_carton']:
                item['units_per_carton'] = excel_vals['units_per_carton']
                changed = True
            if changed:
                updated += 1
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[Sync] Updated {updated} items in {JSON_PATH} from Excel.")

if __name__ == "__main__":
    sync_excel_to_json() 