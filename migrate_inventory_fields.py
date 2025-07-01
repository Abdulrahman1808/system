import os
import json
import openpyxl

EXCEL_PATH = os.path.join('excel_data', 'hookah_inventory.xlsx')
JSON_PATH = os.path.join('mongodb_data', 'inventory.json')

# --- Excel Migration ---
def migrate_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        return
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    changed_rows = 0
    for row in ws.iter_rows(min_row=2):
        # Get indices
        q_idx = headers.get('quantity')
        r_idx = headers.get('retail_quantity')
        c_idx = headers.get('carton_count')
        u_idx = headers.get('units_per_carton')
        # Only migrate if source exists and target is empty or zero
        if q_idx and c_idx and row[q_idx-1].value and (not row[c_idx-1].value or row[c_idx-1].value == 0):
            row[c_idx-1].value = row[q_idx-1].value
            changed_rows += 1
        if r_idx and u_idx and row[r_idx-1].value and (not row[u_idx-1].value or row[u_idx-1].value == 0):
            row[u_idx-1].value = row[r_idx-1].value
            changed_rows += 1
    wb.save(EXCEL_PATH)
    print(f"[Excel] Migrated {changed_rows} fields in {EXCEL_PATH}")

# --- JSON (MongoDB) Migration ---
def migrate_json():
    if not os.path.exists(JSON_PATH):
        print(f"JSON file not found: {JSON_PATH}")
        return
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    changed = 0
    for item in data:
        if 'quantity' in item and ('carton_count' not in item or not item['carton_count']):
            item['carton_count'] = item['quantity']
            changed += 1
        if 'retail_quantity' in item and ('units_per_carton' not in item or not item['units_per_carton']):
            item['units_per_carton'] = item['retail_quantity']
            changed += 1
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[JSON] Migrated {changed} fields in {JSON_PATH}")

if __name__ == "__main__":
    migrate_excel()
    migrate_json()
    print("Migration completed safely.") 